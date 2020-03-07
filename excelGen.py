from bs4 import BeautifulSoup
import os
from openpyxl import load_workbook


def clean(params):
    i = [div.get_text() for div in params].index('Female Ph.D. graduates')
    j = [div.get_text() for div in params].index('Full-time faculty (tenured or tenure-track)')
    return params[:i+1] + params[j:]

def write(params, uni):
    wb = load_workbook('../Unis crawled.xlsx')
    ws = wb.active
    ws.append([uni, ws.max_row] + params)
    wb.save('../Unis crawled.xlsx')


os.chdir('./SavedHTML/')
for index, uni in enumerate(sorted(filter(os.path.isfile, os.listdir('.')), key=os.path.getmtime)):
    if index == 0:
        params = clean(BeautifulSoup(open(uni, 'r', encoding='utf-8').read(), 'html.parser').find_all("div", class_="DataField__Title-s13u9bdi-0"))
        params = [div.get_text() for div in params]
        write(params, 'University')
    params = clean(BeautifulSoup(open(uni, 'r', encoding='utf-8').read(), 'html.parser').find_all("div", class_="DataField__Title-s13u9bdi-0"))
    params = [div.next_sibling.get_text() for div in params]
    write(params, uni)







